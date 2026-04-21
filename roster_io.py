# -*- coding: utf-8 -*-
from __future__ import annotations

from copy import copy
import re
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from roster_logic import (
    DAY_NAMES,
    TITLE_SUFFIX,
    is_duplicate_location,
    normalize_name_text,
    normalize_text,
    parse_title,
)

UNIFORM_FONT_NAME = "Calibri"
UNIFORM_FONT_SIZE = 9
UNIFORM_ROW_HEIGHT = 22
UNIFORM_COLUMN_WIDTH = 18
UNIFORM_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=False)
TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=False)
SIGNATURE_ALIGNMENT = Alignment(horizontal="right", vertical="center", wrap_text=False)
UNIFORM_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
DAY_COLUMN_COUNT = len(DAY_NAMES)
TABLE_COLUMN_COUNT = 1 + DAY_COLUMN_COUNT


def _find_week_title_rows(ws) -> List[Tuple[int, str]]:
    rows = []
    max_row = ws.max_row or 1
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and normalize_text(TITLE_SUFFIX) in normalize_text(val):
            rows.append((r, val))
    return rows


def _detect_day_columns(ws, header_row: int) -> Optional[Dict[str, int]]:
    day_map = {}
    max_col = max(ws.max_column, TABLE_COLUMN_COUNT)
    for c in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=c).value
        if isinstance(val, str):
            v = normalize_text(val)
            if v in DAY_NAMES:
                day_map[v] = c
    if len(day_map) == DAY_COLUMN_COUNT:
        return day_map
    return None


def _extract_colon_value(value: str) -> str:
    if ":" in value:
        return value.split(":", 1)[1].strip()
    return value.strip()


def _extract_school_name_from_text(text: str) -> str:
    clean_text = normalize_name_text(text)
    if not clean_text:
        return ""

    norm = normalize_text(clean_text)
    if norm.startswith("OKUL"):
        return _extract_colon_value(clean_text)
    if normalize_text(TITLE_SUFFIX) in norm:
        return ""
    if norm.startswith("MUDUR") or norm.startswith("MÜDÜR"):
        return ""
    if norm == normalize_text("NOBET YERI"):
        return ""
    if norm in DAY_NAMES:
        return ""

    # New format: school row contains only school name.
    return clean_text


def _find_global_school_name(ws) -> str:
    title_rows = _find_week_title_rows(ws)
    if not title_rows:
        return ""

    first_title_row = title_rows[0][0]
    for row_idx in range(1, first_title_row):
        val = ws.cell(row=row_idx, column=1).value
        if not isinstance(val, str):
            continue
        school_name = _extract_school_name_from_text(val)
        if school_name:
            return school_name
    return ""


def _parse_school_name_from_title_row(ws, title_row: int) -> str:
    if title_row > 1:
        val = ws.cell(row=title_row - 1, column=1).value
        if isinstance(val, str):
            school_name = _extract_school_name_from_text(val)
            if school_name:
                return school_name

    # Multi-week exports may place school name once at sheet top.
    return _find_global_school_name(ws)


def _parse_principal_name_from_signature_row(ws, row_idx: int) -> str:
    for col in (8, 7, 6, 1):
        val = ws.cell(row=row_idx, column=col).value
        if not isinstance(val, str):
            continue
        norm = normalize_text(val)
        if norm.startswith("MUDUR") or norm.startswith("MÜDÜR"):
            return _extract_colon_value(val)
    return ""


def _read_week_block(ws, title_row: int, default_year: int) -> Dict:
    title = ws.cell(row=title_row, column=1).value
    if not isinstance(title, str):
        raise ValueError("Hafta basligi metin degil.")

    header_row = title_row + 1
    day_cols = _detect_day_columns(ws, header_row)
    if not day_cols:
        raise ValueError("Gun basliklari bulunamadi veya eksik.")

    start_date, end_date, _ = parse_title(title, default_year)
    school_name = _parse_school_name_from_title_row(ws, title_row)

    locations: List[str] = []
    roster: List[List[str]] = []

    r = header_row + 1
    max_row = ws.max_row or r
    while r <= max_row:
        location_val = ws.cell(row=r, column=1).value
        location_text = "" if location_val is None else str(location_val).strip()

        row_vals: List[str] = []
        all_empty = True
        for day in DAY_NAMES:
            c = day_cols[day]
            val = ws.cell(row=r, column=c).value
            text = "" if val is None else str(val).strip()
            row_vals.append(text)
            if text:
                all_empty = False

        if not location_text and all_empty:
            break

        if not location_text:
            if not locations:
                break
            location_text = locations[-1]

        locations.append(location_text)
        roster.append(row_vals)
        r += 1

    principal_name = _parse_principal_name_from_signature_row(ws, r)

    return {
        "title": title,
        "start_date": start_date,
        "end_date": end_date,
        "school_name": school_name,
        "principal_name": principal_name,
        "locations": locations,
        "roster": roster,
    }


def load_last_week(path: str, default_year: int) -> Dict:
    wb = load_workbook(path)
    try:
        ws = wb.active
        title_rows = _find_week_title_rows(ws)
        if not title_rows:
            raise ValueError("A sutununda hafta blogu bulunamadi.")
        last_title_row, _ = title_rows[-1]
        return _read_week_block(ws, last_title_row, default_year)
    finally:
        wb.close()


def _find_last_block_end(ws) -> int:
    title_rows = _find_week_title_rows(ws)
    if not title_rows:
        return 0
    last_title_row, _ = title_rows[-1]
    header_row = last_title_row + 1
    day_cols = _detect_day_columns(ws, header_row)
    if not day_cols:
        return header_row

    r = header_row + 1
    max_row = ws.max_row or r
    while r <= max_row:
        location_val = ws.cell(row=r, column=1).value
        location_text = "" if location_val is None else str(location_val).strip()

        all_empty = True
        for day in DAY_NAMES:
            c = day_cols[day]
            val = ws.cell(row=r, column=c).value
            if val not in (None, ""):
                all_empty = False

        if not location_text and all_empty:
            break

        r += 1

    return r - 1


def _apply_duplicate_location_pair_excel_merge(ws, top_row: int, bottom_row: int):
    top_loc = str(ws.cell(row=top_row, column=1).value or "").strip()
    bottom_loc = str(ws.cell(row=bottom_row, column=1).value or "").strip()
    if not is_duplicate_location(top_loc, bottom_loc):
        return

    ws.merge_cells(start_row=top_row, start_column=1, end_row=bottom_row, end_column=1)
    ws.cell(row=top_row, column=1, value=top_loc or bottom_loc)
    ws.cell(row=bottom_row, column=1, value=None)

    # A=location, B..F=weekdays
    for col in range(2, TABLE_COLUMN_COUNT + 1):
        top_name = str(ws.cell(row=top_row, column=col).value or "").strip()
        bottom_name = str(ws.cell(row=bottom_row, column=col).value or "").strip()
        if top_name and bottom_name and normalize_text(top_name) != normalize_text(bottom_name):
            # Keep both rows visible when both are filled.
            continue
        chosen = top_name or bottom_name or None
        ws.cell(row=top_row, column=col, value=chosen)
        ws.cell(row=bottom_row, column=col, value=None)
        ws.merge_cells(start_row=top_row, start_column=col, end_row=bottom_row, end_column=col)


def _apply_duplicate_location_excel_merge_rules(ws, first_row: int, last_row: int):
    row = first_row
    while row < last_row:
        top_loc = str(ws.cell(row=row, column=1).value or "").strip()
        bottom_loc = str(ws.cell(row=row + 1, column=1).value or "").strip()
        if is_duplicate_location(top_loc, bottom_loc):
            _apply_duplicate_location_pair_excel_merge(ws, row, row + 1)
            row += 2
        else:
            row += 1


def _apply_basic_formatting(ws, uniform_font_size: int):
    max_row = max(ws.max_row or 0, 1)
    max_col = TABLE_COLUMN_COUNT

    for col_idx in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = UNIFORM_COLUMN_WIDTH

    for row_idx in range(1, max_row + 1):
        ws.row_dimensions[row_idx].height = UNIFORM_ROW_HEIGHT
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            is_bold = bool(cell.font and cell.font.bold)
            font = copy(cell.font)
            font.name = UNIFORM_FONT_NAME
            font.size = uniform_font_size
            font.bold = is_bold
            cell.font = font
            alignment = copy(UNIFORM_ALIGNMENT)
            alignment.wrap_text = isinstance(cell.value, str) and ("\n" in cell.value)
            cell.alignment = alignment
            cell.border = UNIFORM_BORDER


def _write_school_line(ws, row_idx: int, school_name: str, font_size: int):
    school_clean = normalize_name_text(school_name)
    if not school_clean:
        return

    ws.cell(row=row_idx, column=1, value=school_clean.upper())
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=TABLE_COLUMN_COUNT)
    ws.cell(row=row_idx, column=1).font = Font(name=UNIFORM_FONT_NAME, bold=True, size=font_size)
    ws.cell(row=row_idx, column=1).alignment = TITLE_ALIGNMENT


def _write_principal_line(ws, row_idx: int, principal_name: str, font_size: int):
    principal_clean = normalize_name_text(principal_name)
    signature_text = f"Müdür : {principal_clean}" if principal_clean else "Müdür :"
    sig_cell = ws.cell(row=row_idx, column=8, value=signature_text)
    sig_cell.font = Font(name=UNIFORM_FONT_NAME, bold=True, size=font_size)
    sig_cell.alignment = SIGNATURE_ALIGNMENT
    ws.row_dimensions[row_idx].height = UNIFORM_ROW_HEIGHT


def _first_non_empty_week_field(weeks: List[Dict], field_name: str) -> str:
    for week in weeks:
        value = normalize_name_text(str((week or {}).get(field_name, "") or ""))
        if value:
            return value
    return ""


def append_week_block(
    ws,
    start_row: int,
    title: str,
    locations: List[str],
    roster: List[List[str]],
    uniform_font_size: int,
    school_name: str = "",
    principal_name: str = "",
    *,
    include_school_row: bool = True,
    include_signature_row: bool = True,
) -> int:
    bold = Font(name=UNIFORM_FONT_NAME, bold=True, size=uniform_font_size)

    row_cursor = start_row
    if include_school_row:
        school_clean = normalize_name_text(school_name)
        if school_clean:
            _write_school_line(ws, row_cursor, school_clean, uniform_font_size)
            row_cursor += 1

    title_row = row_cursor
    ws.cell(row=title_row, column=1, value=title)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=TABLE_COLUMN_COUNT)
    ws.cell(row=title_row, column=1).font = bold
    ws.cell(row=title_row, column=1).alignment = TITLE_ALIGNMENT

    header_row = title_row + 1
    headers = ["NOBET YERI"] + DAY_NAMES
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=header)
        cell.font = bold
        cell.fill = HEADER_FILL

    for idx, location in enumerate(locations):
        r = header_row + 1 + idx
        ws.cell(row=r, column=1, value=location)
        row_vals = roster[idx] if idx < len(roster) else [""] * DAY_COLUMN_COUNT
        for c, val in enumerate(row_vals, start=2):
            full_name = normalize_name_text(str(val or ""))
            ws.cell(row=r, column=c, value=full_name)

    if len(locations) >= 2:
        first_body_row = header_row + 1
        last_body_row = header_row + len(locations)
        _apply_duplicate_location_excel_merge_rules(ws, first_body_row, last_body_row)

    _apply_basic_formatting(ws, uniform_font_size)

    last_table_row = header_row + len(locations)
    if include_signature_row:
        signature_row = last_table_row + 1
        _write_principal_line(ws, signature_row, principal_name, uniform_font_size)
        return signature_row

    return last_table_row


def write_week_to_excel(
    source_path: Optional[str],
    title: str,
    locations: List[str],
    roster: List[List[str]],
    output_path: str,
    school_name: str = "",
    principal_name: str = "",
):
    wb = load_workbook(source_path) if source_path else Workbook()
    try:
        ws = wb.active

        last_end = _find_last_block_end(ws)
        start_row = 1 if last_end == 0 else last_end + 2

        append_week_block(
            ws,
            start_row,
            title,
            locations,
            roster,
            UNIFORM_FONT_SIZE,
            school_name=school_name,
            principal_name=principal_name,
        )
        wb.save(output_path)
    finally:
        wb.close()


def write_weeks_to_excel(
    source_path: Optional[str],
    weeks: List[Dict],
    output_path: str,
):
    wb = load_workbook(source_path) if source_path else Workbook()
    try:
        ws = wb.active

        valid_weeks = [week for week in weeks if str((week or {}).get("title", "") or "").strip()]

        last_end = _find_last_block_end(ws)
        start_row = 1 if last_end == 0 else last_end + 2

        batch_school_name = _first_non_empty_week_field(valid_weeks, "school_name")
        batch_principal_name = _first_non_empty_week_field(valid_weeks, "principal_name")

        if batch_school_name:
            _write_school_line(ws, start_row, batch_school_name, UNIFORM_FONT_SIZE)
            start_row += 1

        for week in valid_weeks:
            title = str((week or {}).get("title", "") or "")
            locations = list((week or {}).get("locations", []) or [])
            roster = list((week or {}).get("roster", []) or [])

            block_end = append_week_block(
                ws,
                start_row,
                title,
                locations,
                roster,
                UNIFORM_FONT_SIZE,
                school_name="",
                principal_name="",
                include_school_row=False,
                include_signature_row=False,
            )
            start_row = block_end + 1

        if valid_weeks:
            _write_principal_line(ws, start_row, batch_principal_name, UNIFORM_FONT_SIZE)

        wb.save(output_path)
    finally:
        wb.close()
